import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from .models import CustomUser, Vehicle, Driver, Technician, Party, Product, Purchase, JobCard
from .filters import *
from openpyxl.styles import Font 
from datetime import date


from django.utils.timezone import localtime
from django.db.models import Sum, Count, Q, F


def export_filtered_job_cards(request):
    # Apply the same filter as in the job_card_list view
    queryset = JobCard.objects.all().order_by('id') 
    filter = JobCardFilter(request.GET, queryset=queryset)
    filtered_job_cards = filter.qs  # Filtered queryset

    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Cards"

    # Define the header row for the Excel sheet
    headers = ['Job Card Number', 'Technician', 'Driver', 'Created Date',  'Reported Defect', 'Completed Action', 'Completed Date', 'Vehicle', 'Status',  'Total Product Amount', "Labour Amount","Grand Total Amount" ]
    ws.append(headers)

    # Add data rows to the Excel sheet for each filtered job card
    for job_card in filtered_job_cards:
        # Convert the `date` field to naive datetime
        create_datetime = localtime(job_card.date).replace(tzinfo=None) if job_card.date else ""
        completed_datetime = localtime(job_card.completed_date).replace(tzinfo=None) if job_card.completed_date else ""
        
        items = JobCardItem.objects.filter(job_card=job_card).order_by('-id')
        total_cost = items.aggregate(Sum('total_cost'))['total_cost__sum']
        total_cost = int(total_cost) if total_cost is not None else 0
        labour_cost=int(job_card.labour_cost) if job_card.labour_cost is not None else 0
        grand_total_cost=int(total_cost+labour_cost)
        row = [
            job_card.job_card_number,
            job_card.technician.technician_name if job_card.technician else "",
            job_card.driver.driver_name if job_card.driver else "",
            create_datetime,  
            job_card.reported_defect,
            job_card.completed_action,
            completed_datetime,
            job_card.vehicle.vehicle_name if job_card.vehicle else "",
            job_card.status,
            total_cost,
            labour_cost,
            grand_total_cost,
        ]
        ws.append(row)

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=job_cards.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response


def export_purchase_data(request):
    # Apply the same filter as in the job_card_list view
    queryset = Purchase.objects.all().order_by('id') 
    filter = PurchaseFilter(request.GET, queryset=queryset)
    filtered_data = filter.qs  # Filtered queryset

    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Data"

    # Define the header row for the Excel sheet
    headers = ['Bill Number', 'Supplier Name', 'Bill Type', 'Bill Date',  'Total Amount']
    ws.append(headers)

    # Add data rows to the Excel sheet for each filtered job card
    for data in filtered_data:
        # Convert the `date` field to naive datetime
        # date = localtime(data.bill_date).replace(tzinfo=None) if data.bill_date else ""
 
        row = [
            data.bill_no,
            data.supplier_name if data.supplier_name else "",
            data.bill_type,
            data.bill_date,
            data.total_cost,
        ]
        ws.append(row)

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Purchase.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response




def export_product_data(request):
    # Fetch all products or apply filters if needed
    queryset = Product.objects.all().order_by('id')

    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Data"

    # Define the header row for the Excel sheet
    headers = [
        'Product Code', 'Product Name', 'Model', 'Description', 
        'Sale Price', 'Minimum Stock Alert', 'Available Stock'
    ]
    ws.append(headers)

    # Add data rows to the Excel sheet for each product
    for product in queryset:
        row = [
            product.product_code,
            product.product_name,
            product.model.model_name if product.model else "N/A",  # Access related `model_name`
            product.description if product.description else "",
            product.sale_price if product.sale_price else 0.00,
            product.minimum_stock_alert,
            product.available_stock,
        ]
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ProductData.xlsx'
    wb.save(response)
    return response





def export_vehicle_data(request):
    # Fetch all vehicles from the database
    queryset = Vehicle.objects.all().order_by('id')

    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicle Data"

    # Define the header row for the Excel sheet
    headers = ['ID', 'Vehicle Name', 'Vehicle Number']
    ws.append(headers)

    # Add data rows to the Excel sheet for each vehicle
    for vehicle in queryset:
        row = [
            vehicle.id,
            vehicle.vehicle_name,
            vehicle.vehicle_number,
        ]
        ws.append(row)

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=VehicleData.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response


from django.conf import settings
def export_driver_data(request):
    # Fetch all drivers from the database
    queryset = Driver.objects.all().order_by('id')
    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Driver Data"

    # Define the header row for the Excel sheet
    headers = [
        'ID','EMP ID', 'Driver Name', 'License Number', 'Adhaar Number', 
        'Mobile Number', 'Alternate Mobile Number', 'Address', 
        'Date of Birth', 'Date Joined', 'Adhaar Card Photo', 
        'PAN Card Photo', 'Driving License Photo', 'Profile Photo'
    ]

    ws.append(headers)
    # Add data rows to the Excel sheet for each driver
    for driver in queryset:
        row = [
            driver.id,
            str(driver.user.emp_id) if str(driver.user.emp_id) else "",
            driver.driver_name,
            driver.license_number if driver.license_number else "",
            driver.adhaar_number if driver.adhaar_number else "",
            driver.mobile_number if driver.mobile_number else "",
            driver.alternate_mobile_number if driver.alternate_mobile_number else "",
            driver.address if driver.address else "",
            driver.date_of_birth if driver.date_of_birth else "",
            driver.date_joined if driver.date_joined else "",
            request.build_absolute_uri(driver.adhaar_card_photo.url) if driver.adhaar_card_photo else "",
            request.build_absolute_uri(driver.pan_card_photo.url) if driver.pan_card_photo else "",
            request.build_absolute_uri(driver.driving_license_photo.url) if driver.driving_license_photo else "",
            request.build_absolute_uri(driver.profile_photo.url) if driver.profile_photo else "",
        ]
        ws.append(row)

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=DriverData.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response




def export_technician_data(request):
    # Fetch all technicians from the database
    queryset = Technician.objects.all().order_by('id')

    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Technician Data"

    # Define the header row for the Excel sheet
    headers = [
        'EMP ID', 'Technician Name', 'Aadhaar Number', 'Mobile Number', 
        'Alternate Mobile Number', 'Email Address', 'Address', 
        'Date of Birth', 'Date Joined', 'PAN Card', 'Aadhaar Card', 
        'Profile Photo', 'Additional Docs'
    ]
    ws.append(headers)

    # Add data rows to the Excel sheet for each technician
    for technician in queryset:
        row = [
            str(technician.emp_id)  if str(technician.email) else "",
            technician.technician_name,
            technician.adhaar_number,
            technician.mobile_number,
            technician.alternate_mobile_number if technician.alternate_mobile_number else "",
            technician.email if technician.email else "",
            technician.address if technician.address else "",
            technician.date_of_birth if technician.date_of_birth else "",
            technician.date_joined if technician.date_joined else "",
            request.build_absolute_uri(technician.pan_card.url) if technician.pan_card else "",
            request.build_absolute_uri(technician.adhaar_card.url) if technician.adhaar_card else "",
            request.build_absolute_uri(technician.profile_photo.url) if technician.profile_photo else "",
            request.build_absolute_uri(technician.additional_docs.url) if technician.additional_docs else "",
        ]
        ws.append(row)

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=TechnicianData.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response




def export_party_data(request):
    # Fetch all parties from the database
    queryset = Party.objects.all().order_by('id')

    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Party Data"

    # Define the header row for the Excel sheet
    headers = [
        'ID', 'Business Name', 'Customer Name', 'GST Number', 
        'Mobile Number', 'Alternate Mobile Number', 'Address', 
        'Document 1', 'Document 2', 'Document 3'
    ]
    ws.append(headers)

    # Add data rows to the Excel sheet for each party
    for party in queryset:
        row = [
            party.id,
            party.business_name,
            party.customer_name,
            party.gst_number if party.gst_number else "",
            party.mobile_number,
            party.alternate_mobile_number if party.alternate_mobile_number else "",
            party.address if party.address else "",
            request.build_absolute_uri(party.document1.url) if party.document1 else "",
            request.build_absolute_uri(party.document2.url) if party.document2 else "",
            request.build_absolute_uri(party.document3.url) if party.document3 else "",
        ]
        ws.append(row)

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=PartyData.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response



 

def export_policy_data(request):
    # Fetch policies and calculate remaining days
    policies = Policy.objects.select_related('vehicle').order_by('due_date')
    policy_data = [
        {
            'policy_number': policy.policy_number,
            'vehicle_name': policy.vehicle.vehicle_name,
            'vehicle_number': policy.vehicle.vehicle_number,
            'due_date': policy.due_date,
            'remaining_days': (policy.due_date - date.today()).days
        }
        for policy in policies
    ]

    # Create a new Workbook and add a sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Policy Data"

    # Define the header row for the Excel sheet
    headers = ['Policy Number', 'Vehicle Name', 'Vehicle Number', 'Due Date', 'Remaining Days']
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font

    # Add data rows to the Excel sheet for each policy
    for row_num, data in enumerate(policy_data, 2):
        ws.cell(row=row_num, column=1, value=data['policy_number'])  # Policy Number
        ws.cell(row=row_num, column=2, value=data['vehicle_name'])  # Vehicle Name
        ws.cell(row=row_num, column=3, value=data['vehicle_number'])  # Vehicle Number
        ws.cell(row=row_num, column=4, value=data['due_date'].strftime('%Y-%m-%d'))  # Due Date
        ws.cell(row=row_num, column=5, value=data['remaining_days'])  # Remaining Days

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=PolicyData.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response



def export_emi_data(request):
    # Fetch all EMI records from the database
    queryset = EMI.objects.select_related('vehicle').all().order_by('id')
    filter = EMIFilter(request.GET, queryset=queryset)
    filtered_data = filter.qs  # Filtered queryset

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EMI Data"

    headers = [
        'Vehicle No', 'Loan Amount', 'Total Installments', 
        'Paid Installments', 'Remaining Installments', 
        'Next Due Date', 'Days Remaining for Due', 
        'Status', 'Frequency'
    ]
    ws.append(headers)

    # Add data rows to the Excel sheet for each EMI record
    for emi in filtered_data:
        # Calculate remaining installments and days remaining for the next due date
        remaining_installments = emi.remaining_installments
        days_remaining = (emi.next_due_date - date.today()).days if emi.next_due_date else None

        row = [
            emi.vehicle.vehicle_number if emi.vehicle else "N/A",  # Vehicle No
            emi.loan_amount,  # Loan Amount
            emi.total_installments,  # Total Installments
            emi.paid_installments,  # Paid Installments
            remaining_installments,  # Remaining Installments
            emi.next_due_date if emi.next_due_date else "N/A",  # Next Due Date
            days_remaining if days_remaining is not None else "N/A",  # Days Remaining
            emi.status,  # Status
            emi.frequency,  # Frequency
        ]
        ws.append(row)

    # Create an HTTP response with the Excel file content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EMIData.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response



 
from .filters import VehicleFilterForFinance
def export_vehicle_for_finance(request):
    # Instantiate the filter form
    filter_form = VehicleFilterForFinance(request.GET or None)
    queryset = Vehicle.objects.all()

    # Initialize filters
    start_date = filter_form.cleaned_data.get('start_date') if filter_form.is_valid() else None
    end_date = filter_form.cleaned_data.get('end_date') if filter_form.is_valid() else None

    # Prepare the filtered data
    vehicles_data = []
    for vehicle in queryset:
        # Get related data for each vehicle
        policies = vehicle.policies.all()
        emis = vehicle.emis.all()
        insurance_tax = vehicle.insurancetaxdue_set.first()  # Assuming one insurance/tax record

        # Extract relevant data
        policy_due_date = policies[0].due_date if policies else None
        policy_number = policies[0].policy_number if policies else None
        policy_file = policies[0].policy_file.url if policies else None

        finance_bank = emis[0].finance_bank.bank_name if emis else None
        loan_account_no = emis[0].loan_account_no if emis else None
        emi_due_date = emis[0].next_due_date if emis else None
        emi_frequency = emis[0].frequency if emis else None
        emi_status = emis[0].status if emis else None
        emi_file = emis[0].file.url if emis else None
        total_installments = emis[0].total_installments if emis else None
        paid_installments = emis[0].paid_installments if emis else None
        remaining_installments = emis[0].remaining_installments if emis else None

        insurance_bank = insurance_tax.insurance_bank.bank_name if insurance_tax.insurance_bank else None
        insurance_amount = insurance_tax.insurance_tax.insurance_amount if insurance_tax else None
        insurance_due_date = insurance_tax.insurance_due_date if insurance_tax else None
        insurance_amount = insurance_tax.insurance_amount if insurance_tax else None
        tax_due_date = insurance_tax.tax_due_date if insurance_tax else None
        tax_amount = insurance_tax.tax_amount if insurance_tax else None
        fitness_due_date = insurance_tax.fitness_due_date if insurance_tax else None
        permit_due_date = insurance_tax.permit_due_date if insurance_tax else None
        puc_due_date = insurance_tax.puc_due_date if insurance_tax else None

        # Apply date filtering
        if start_date and not end_date:
            if not (
                (policy_due_date and policy_due_date >= start_date) or
                (emi_due_date and emi_due_date >= start_date) or
                (insurance_due_date and insurance_due_date >= start_date) or
                (tax_due_date and tax_due_date >= start_date) or
                (fitness_due_date and fitness_due_date >= start_date) or
                (permit_due_date and permit_due_date >= start_date) or
                (puc_due_date and puc_due_date >= start_date)
            ):
                continue
        elif end_date and not start_date:
            if not (
                (policy_due_date and policy_due_date <= end_date) or
                (emi_due_date and emi_due_date <= end_date) or
                (insurance_due_date and insurance_due_date <= end_date) or
                (tax_due_date and tax_due_date <= end_date) or
                (fitness_due_date and fitness_due_date <= end_date) or
                (permit_due_date and permit_due_date <= end_date) or
                (puc_due_date and puc_due_date <= end_date)
            ):
                continue
        elif start_date and end_date:
            if not (
                (policy_due_date and start_date <= policy_due_date <= end_date) or
                (emi_due_date and start_date <= emi_due_date <= end_date) or
                (insurance_due_date and start_date <= insurance_due_date <= end_date) or
                (tax_due_date and start_date <= tax_due_date <= end_date) or
                (fitness_due_date and start_date <= fitness_due_date <= end_date) or
                (permit_due_date and start_date <= permit_due_date <= end_date) or
                (puc_due_date and start_date <= puc_due_date <= end_date)
            ):
                continue

        # Add vehicle info to the export data
        vehicles_data.append({
            'vehicle_number': vehicle.vehicle_number,
            'policy_due_date': policy_due_date,
            'policy_number': policy_number,
            'policy_file': policy_file,
            'finance_bank': finance_bank,
            'loan_account_no': loan_account_no,
            'emi_due_date': emi_due_date,
            'emi_frequency': emi_frequency,
            'emi_status': emi_status,
            'emi_file': emi_file,
            'total_installments': total_installments,
            'paid_installments': paid_installments,
            'remaining_installments': remaining_installments,
            'insurance_bank': insurance_bank,
            'insurance_amount': insurance_amount,
            'insurance_due_date': insurance_due_date,
            'insurance_amount': insurance_amount,
            'tax_due_date': tax_due_date,
            'tax_amount': tax_amount,
            'fitness_due_date': fitness_due_date,
            'permit_due_date': permit_due_date,
            'puc_due_date': puc_due_date,
        })

    # Create an Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Vehicles'

    # Add header row with bold font
    header = [
        'Vehicle Number', 'Policy Due Date', 'Policy Number', 'Policy File',
        'EMI Finance Bank','Loan Account No','EMI Due Date', 'EMI Frequency', 'EMI Status', 'EMI File', 'Total Installments','Paid Installments','Remaining Installments',
        'Insurance Bank','Insurance Amount','Insurance Due Date', 'Insurance Amount', 'Tax Due Date', 'Tax Amount',
        'Fitness Due Date', 'Permit Due Date', 'PUC Due Date'
    ]
    for col_num, column_title in enumerate(header, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)

    # Add vehicle data to rows
    for row_num, vehicle in enumerate(vehicles_data, 2):
        for col_num, field in enumerate(header, 1):
            sheet.cell(row=row_num, column=col_num).value = vehicle.get(field.lower().replace(' ', '_'))

    # Create HTTP response for download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=vehicles.xlsx'
    workbook.save(response)

    return response


