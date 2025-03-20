from django.shortcuts import render,redirect,get_object_or_404
from .models import *
from django.contrib import messages
import openpyxl
from .forms import *
from django.http import JsonResponse
from django.db.models import Q
from django.contrib.auth import login as authlogin, authenticate,logout as DeleteSession

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login as auth_login
from django.contrib import messages
from django.db.models import Sum, Count, Q, F
from django.contrib.auth.hashers import make_password

from django.core.cache import cache

from django.db.models.signals import post_save, post_delete
from django.dispatch import receiver
from django.forms.models import model_to_dict
from django.core.cache import cache
from .models import *
from django.core.paginator import Paginator
from .filters import *
from functools import wraps

from ERP_Workshop.filters import ProductFilter
from ERP_Workshop.forms import ProductForm

def admin_required(function):
    @wraps(function)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/login')  # Redirect to login if not authenticated
        if not (request.user.is_admin or request.user.is_superuser):
            return redirect('/login')  # Redirect to login if user is not admin or superuser
        return function(request, *args, **kwargs)
    return _wrapped_view
 

def send_email_in_background(email_message):
    try:
        email_message.send()
    except Exception as e:
        print(f"Error sending email: {e}")


@receiver(post_save, sender=Driver)
def update_driver_cache_on_save(sender, instance, **kwargs):
    # Retrieve all drivers and include all fields
    drivers = Driver.objects.select_related('user').all()
    
    # Use model_to_dict to serialize the driver model and its related data
    from django.forms.models import model_to_dict
    driver_data = [model_to_dict(driver) for driver in drivers]

    # Update the cache with all driver data, including related CustomUser fields
    cache.set('cache_drivers', driver_data, timeout=None)
    print("Driver cache updated on save")

@receiver(post_delete, sender=Driver)
def update_driver_cache_on_delete(sender, instance, **kwargs):
    # Retrieve all drivers and include all fields
    drivers = Driver.objects.select_related('user').all()

    # Serialize the driver model data into dictionary format
    from django.forms.models import model_to_dict
    driver_data = [model_to_dict(driver) for driver in drivers]

    # Update the cache to reflect the deletion
    cache.set('cache_drivers', driver_data, timeout=None)
    print("Driver cache updated on delete")


@receiver(post_save, sender=Technician)
def update_technician_cache_on_save(sender, instance, **kwargs):
    # Retrieve all technicians and include all fields
    technicians = Technician.objects.all()
    
    # Use model_to_dict to serialize the technician model
    from django.forms.models import model_to_dict
    technician_data = [model_to_dict(technician) for technician in technicians]

    # Update cache with the latest technician data
    cache.set('cache_technicians', technician_data, timeout=None)
    print("Technician cache updated on save")

@receiver(post_delete, sender=Technician)
def update_technician_cache_on_delete(sender, instance, **kwargs):
    # Retrieve all technicians and include all fields
    technicians = Technician.objects.all()

    # Serialize the technician model data into dictionary format
    from django.forms.models import model_to_dict
    technician_data = [model_to_dict(technician) for technician in technicians]

    # Update the cache to reflect the deletion
    cache.set('cache_technicians', technician_data, timeout=None)
    print("Technician cache updated on delete")




@receiver(post_save, sender=Party)
def update_party_cache_on_save(sender, instance, **kwargs):
    # Retrieve all partys data and serialize it if needed
    party = Party.objects.all()
    # Serialize the technician model data into dictionary format
    from django.forms.models import model_to_dict
    party_data = [model_to_dict(party) for party in party]
    print("Party cache updated on save")

@receiver(post_delete, sender=Party)
def update_party_cache_on_delete(sender, instance, **kwargs):
    # Retrieve all partys data after delete
    partys = []
    for partys in Party.objects.all(): 
        party_data = model_to_dict(partys)  # Convert technician to dictionary
        partys.append(party_data)
    cache.set('cache_technicians', partys, timeout=None)

     
def reload_all_caches(request):
    # Reload Product cache (storing all data as dictionaries)
    products = []
    for product in Product.objects.all().order_by('-id'):
        product_data = model_to_dict(product)  # Convert product to dictionary
        products.append(product_data)
    cache.set('cache_products', products, timeout=None)
    print("Product cache reloaded")

    # Reload Driver cache (storing all data as dictionaries)
    drivers = []
    for driver in Driver.objects.select_related('user').all():  # Add related fields
        driver_data = model_to_dict(driver)  # Convert driver to dictionary
        # If you want to include related fields more explicitly, you can manually add them here
        driver_data['user'] = model_to_dict(driver.user)  # Include related user data
        drivers.append(driver_data)
    cache.set('cache_drivers', drivers, timeout=None)
    print("Driver cache reloaded")

    # Reload Technician cache (storing all data as dictionaries)
    technicians = []
    for technician in Technician.objects.all():
        technician_data = model_to_dict(technician)  # Convert technician to dictionary
        technicians.append(technician_data)
    cache.set('cache_technicians', technicians, timeout=None)
    print("Technician cache reloaded")

    # Reload Party cache (storing all data as dictionaries)
    partys = []
    for party in Party.objects.all():
        party_data = model_to_dict(party)  # Convert party to dictionary
        partys.append(party_data)
    cache.set('cache_partys', partys, timeout=None)
    print("Party cache reloaded")
    return redirect('/admin/dashboard')




def login(request): 
    if request.user.is_authenticated:
        return redirect_user_based_on_role(request, request.user)

    if request.method == 'POST': 
        username = request.POST.get('username', None)
        password = request.POST.get('password', None)
        user = authenticate(request, username=username, password=password)
        if user is not None:
            auth_login(request, user)
            return redirect_user_based_on_role(request, user)
        else:
            messages.error(request, 'Oops...! User does not exist. Please try again.')

    return render(request, 'login.html')



def update_driver_password(request, id):
    driver = get_object_or_404(Driver, id=id)  # Retrieve the Driver instance safely
    user = get_object_or_404(CustomUser, id=driver.user.id)  # Retrieve the associated CustomUser safely
    if request.method == 'POST':
        form = UpdatePasswordForm(request.POST)
        if form.is_valid():
            # Set the new password for the user and save
            new_password = form.cleaned_data['new_password1']
            user.set_password(new_password)
            user.save()
            messages.success(request, 'Password updated successfully.')
    else:
        form = UpdatePasswordForm()
    return render(request, 'admin_update_driver_password.html', {'form': form, 'driver': driver})


def redirect_user_based_on_role(request, user):
    if user.is_superuser or user.is_admin:
        return redirect('/admin/dashboard')
    elif user.is_workshop:
        return redirect('/workshop/dashboard')
    elif user.is_account:
        return redirect('/account/dashboard')
    elif user.is_driver:
        return redirect('/driver/dashboard')
    elif user.is_finance:
        return redirect('/finance/dashboard')
    elif user.is_fuel:
        DeleteSession(request)
        messages.info(request, 'Fuel role not available for web. Try to use Android Application....!')
        return redirect('/login')
    else:
        messages.error(request, 'Unauthorized user role.')
        return redirect('login')

def logout(request):
    DeleteSession(request)
    return redirect('/login')


@admin_required 
def dashboard(request):
    today = date.today()
    thirty_days_date = today + timedelta(days=30)
    two_days_date = today + timedelta(days=2)

    # Fetch active vehicle IDs in one optimized query
    active_vehicle_ids = Vehicle.objects.filter(status='active').values_list('id', flat=True)

    # Pre-filter dues
    policy_dues = Policy.objects.filter(vehicle_id__in=active_vehicle_ids)
    emi_dues = EMI.objects.filter(vehicle_id__in=active_vehicle_ids)
    other_dues = OtherDues.objects.filter(vehicle_id__in=active_vehicle_ids)

    def get_due_counts(queryset, date_field, date_range):
        return queryset.filter(**{f'{date_field}__range': date_range, f'{date_field}__isnull': False}).count()

    thirty_days_counts = sum([
        get_due_counts(policy_dues, 'due_date', [today, thirty_days_date]),
        get_due_counts(emi_dues, 'next_due_date', [today, thirty_days_date]),
        get_due_counts(other_dues, 'tax_due_date', [today, thirty_days_date]),
        get_due_counts(other_dues, 'fitness_due_date', [today, thirty_days_date]),
        get_due_counts(other_dues, 'permit_due_date', [today, thirty_days_date]),
        get_due_counts(other_dues, 'puc_due_date', [today, thirty_days_date])
    ])

    two_days_counts = sum([
        get_due_counts(policy_dues, 'due_date', [today, two_days_date]),
        get_due_counts(emi_dues, 'next_due_date', [today, two_days_date]),
        get_due_counts(other_dues, 'tax_due_date', [today, two_days_date]),
        get_due_counts(other_dues, 'fitness_due_date', [today, two_days_date]),
        get_due_counts(other_dues, 'permit_due_date', [today, two_days_date]),
        get_due_counts(other_dues, 'puc_due_date', [today, two_days_date])
    ])

    expire_dues_counts = sum([
        policy_dues.filter(due_date__lt=today, due_date__isnull=False).count(),
        emi_dues.filter(next_due_date__lt=today, next_due_date__isnull=False).count(),
        other_dues.filter(tax_due_date__lt=today, tax_due_date__isnull=False).count(),
        other_dues.filter(fitness_due_date__lt=today, fitness_due_date__isnull=False).count(),
        other_dues.filter(permit_due_date__lt=today, permit_due_date__isnull=False).count(),
        other_dues.filter(puc_due_date__lt=today, puc_due_date__isnull=False).count()
    ])

    # Aggregate user role counts in a single query
    user_roles = CustomUser.objects.aggregate(
        admin_count=Count('id', filter=Q(is_admin=True)),
        account_count=Count('id', filter=Q(is_account=True)),
        workshop_count=Count('id', filter=Q(is_workshop=True)),
        driver_count=Count('id', filter=Q(is_driver=True))
    )

    user_count = sum(user_roles.values())

    # Product & Vehicle Counts
    product_count = Product.objects.count()
    vehicle_count = len(active_vehicle_ids)  # Avoiding extra count query

    # Aggregate job card counts
    job_card_status_counts = JobCard.objects.values('status').annotate(count=Count('id'))
    status_map = {item['status']: item['count'] for item in job_card_status_counts}

    job_pending_count = status_map.get('pending', 0)
    job_in_progress_count = status_map.get('in_progress', 0)
    job_completed_count = status_map.get('completed', 0)
    job_total_count = sum(status_map.values())

    # Aggregate stock data
    stock = Product.objects.aggregate(
        total_products=Count('id'),
        low_stock=Count('id', filter=Q(available_stock__lt=F('minimum_stock_alert'), available_stock__gt=0)),
        out_of_stock=Count('id', filter=Q(available_stock__lte=0))
    )

    context = {
        'roles': ['Admin', 'Account', 'Workshop', 'Driver'],
        'counts': list(user_roles.values()),
        'user_count': user_count,
        'product_count': product_count,
        'vehicle_count': vehicle_count,
        'job_pending_count': job_pending_count,
        'job_in_progress_count': job_in_progress_count,
        'job_completed_count': job_completed_count,
        'job_total_count': job_total_count,
        'thirty_days_counts': thirty_days_counts,
        'two_days_counts': two_days_counts,
        'expire_dues_counts': expire_dues_counts,
        'low_stock': stock['low_stock'],
        'out_of_stock': stock['out_of_stock'],
        'total_products': stock['total_products'],
    }

    return render(request, "admin_dashboard.html", context)

@admin_required
def enquiry_list(request):
    queryset = Enquiry.objects.order_by('-id')  # No need for `.all()`
    # Apply filtering
    filter = EnquiryFilter(request.GET, queryset=queryset)
    filtered_enquiry = filter.qs
    return render(request, 'admin_enquiry_list.html', {'filtered_enquiry': filtered_enquiry, 'filter': filter})

 
def delete_enquiry(request,id):
    Enquiry.objects.get(id=id).delete()
    return redirect("/admin/enquiry_list")


@admin_required
def vehicle_model_list(request):
    try:
        rec = VehicleModel.objects.all()
        if request.method == 'POST':
            form = VehicleModelForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Model Created Successfully.')
                return redirect('/admin/vehicle_model_list')
            else:
                messages.error(request, 'Form is invalid. Please check your input.')
        else: 
            form = VehicleModelForm()
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
        form = VehicleModelForm()
        rec = []

    return render(request, 'admin_vehicle_model_list.html', {'rec': rec, 'form': form})

 
def vehicle_model_update(request, id):
    try:
        model_instance = get_object_or_404(VehicleModel, id=id)
        if request.method == 'POST':
            form = VehicleModelForm(request.POST, instance=model_instance)
            if form.is_valid():
                form.save()
                messages.success(request, 'Model updated successfully.')
                return redirect('/admin/vehicle_model_list')  # Redirect after successful update
            else:
                messages.error(request, 'Form is invalid. Please check your input.')
        else:
            form = VehicleModelForm(instance=model_instance)
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
        form = None  # Prevents passing an uninitialized variable in case of an error
    return render(request, 'admin_vehicle_model_update.html', {'form': form})

def vehicle_model_delete(request, id):
    try:
        model_instance = get_object_or_404(VehicleModel, id=id)
        model_instance.delete()
        messages.success(request, 'Model deleted successfully.')
    except Exception as e:
        messages.error(request, f'Error : {str(e)}')
    return redirect('/admin/vehicle_model_list')


def notifications(request):
    return render(request, "admin_notifications.html")

def financial_management(request):
    return render(request, "admin_financial_management.html")

def live_status(request):
    return render(request, "admin_live_status.html")
 
@admin_required
def job_card_list(request):
    try:
        queryset = JobCard.objects.order_by('-id') 
        filter = JobCardFilter(request.GET, queryset=queryset)
        filtered_job_cards = filter.qs  # Filtered queryset

        # Pagination
        paginator = Paginator(filtered_job_cards, 20)  # Show 20 job cards per page
        page_number = request.GET.get('page')  # Get the page number from the GET request
        page_obj = paginator.get_page(page_number)  # Get the corresponding page object

        # Include the filter parameters in the pagination context
        filter_params = request.GET.copy()  # Copy the GET parameters
        if 'page' in filter_params:
            del filter_params['page']  # Remove the page parameter if it exists

        # Job card status counts
        pending_count = filtered_job_cards.filter(status="pending").count()
        in_progress_count = filtered_job_cards.filter(status="in_progress").count()
        completed_count = filtered_job_cards.filter(status="completed").count()
        total_count = pending_count + in_progress_count + completed_count

    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        page_obj = None  # Prevents issues if pagination fails
        filter = None
        pending_count = in_progress_count = completed_count = total_count = 0

    return render(request, "admin_job_card_list.html", {
        'job_card': page_obj,  # Pass the paginated object to the template
        'filter': filter,  # Pass the filter object for displaying the form
        'pending_count': pending_count,
        'in_progress_count': in_progress_count,
        'completed_count': completed_count,
        'total_count': total_count
    })


@admin_required
def job_card_item_list(request, id):
    try:
        job_card = get_object_or_404(JobCard, id=id)

        # Fetch all job card items with an optimized query
        items = JobCardItem.objects.filter(job_card=job_card).order_by('-id')

        # Aggregate total cost efficiently
        total_cost = items.aggregate(total_cost=Sum('total_cost'))['total_cost'] or 0

        # Ensure labour_cost is handled safely
        labour_cost = job_card.labour_cost or 0

        # Grand total cost calculation
        grand_total_cost = total_cost + labour_cost

    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        items = []
        total_cost = labour_cost = grand_total_cost = 0
        job_card = None

    return render(request, "admin_job_card_item_list.html", {
        'items': items,
        'job_card': job_card,
        'total_cost': total_cost,
        'labour_cost': labour_cost,
        'grand_total_cost': grand_total_cost
    })


@admin_required
def purchase_list(request):
    try:
        queryset = Purchase.objects.order_by('-id')
        filter = PurchaseFilter(request.GET, queryset=queryset)
        filtered_purchase = filter.qs  # Filtered queryset

        # Compute total count and total amount before pagination (efficiently)
        total_bill_count = filtered_purchase.count()
        total_amount = filtered_purchase.aggregate(total_cost=Sum('total_cost'))['total_cost'] or 0.0

        # Pagination
        paginator = Paginator(filtered_purchase, 20)  # Show 20 purchases per page
        page_number = request.GET.get('page')  # Get the page number from the GET request
        page_obj = paginator.get_page(page_number)  # Get the corresponding page object

        # Preserve filter parameters for pagination
        filter_params = request.GET.copy()
        filter_params.pop('page', None)  # Remove 'page' parameter if it exists

    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        page_obj = []
        total_bill_count = 0
        total_amount = 0.0
        filter_params = {}

    return render(request, "admin_purchase_list.html", {
        'purchase': page_obj,  # Pass the paginated object to the template
        'filter': filter,  # Pass the filter object for displaying the form
        'filter_params': filter_params.urlencode(),  # Pass the filter parameters for pagination
        'total_bill_count': total_bill_count,
        'total_amount': total_amount
    })

from django.db import DatabaseError
 
@admin_required
def purchase_item_list(request, id):
    try:
        purchase = get_object_or_404(Purchase, id=id)

        # Optimize query: fetch related purchase and order results
        items = PurchaseItem.objects.filter(purchase=purchase).order_by('-id').select_related('purchase')

        # Aggregate total amount efficiently
        total_amount = items.aggregate(total_amount=Sum('total_amount'))['total_amount'] or 0

    except DatabaseError as e:  # Catch only database-related errors
        messages.error(request, f"Database error: {str(e)}")
        purchase, items, total_amount = None, [], 0

    return render(request, "admin_purchase_item_list.html", {
        'item': items,
        'purchase': purchase,
        'total_amount': int(total_amount)  # Ensure integer format
    })


@admin_required
def product_list(request): 
    try:
        queryset = Product.objects.select_related("model").order_by("-id")

        # Apply filtering
        filter = ProductFilter(request.GET, queryset=queryset)
        filtered_rec = filter.qs  # Apply filters to the queryset

        # Pagination setup
        paginator = Paginator(filtered_rec, 20)  # Show 20 products per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Preserve filter parameters for pagination
        filter_params = request.GET.copy()
        filter_params.pop('page', None)  # More efficient way to remove 'page' if it exists

        # Optimize stock calculations
        out_of_stock_items = filtered_rec.filter(available_stock__lte=0).values_list("id", flat=True).count()
        available_stock_items = filtered_rec.filter(available_stock__gt=F('minimum_stock_alert')).values_list("id", flat=True).count()
        low_stock_items = filtered_rec.filter(available_stock__lte=F('minimum_stock_alert'), available_stock__gt=0).values_list("id", flat=True).count()

    except DatabaseError as db_err:
        messages.error(request, f"Database error: {db_err}")
        page_obj = []
        out_of_stock_items = available_stock_items = low_stock_items = 0

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        page_obj = []
        out_of_stock_items = available_stock_items = low_stock_items = 0

    return render(request, "admin_product_list.html", {
        'form': ProductForm(),
        'product': page_obj,
        'filter': filter,
        'filter_params': filter_params.urlencode(),
        'out_of_stock_items': out_of_stock_items,
        'available_stock_items': available_stock_items,
        'low_stock_items': low_stock_items,
    })


def report(request):
    return render(request, "admin_report.html")

@admin_required
def user_management(request):
    form=UserRegistrationForm()
    users=CustomUser.objects.filter(is_driver=False,is_superuser=False)
    return render(request, "admin_user_management.html",{'form':form,'users':users})

def create_user(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST,request.FILES)
        if form.is_valid():
            try:
                fm=form.save(commit=False)
                last_emp_id = EMP_ID.objects.order_by('-id').first()

                if last_emp_id:
                    last_number = int(last_emp_id.emp_id.split('-')[1])
                    new_emp_id = EMP_ID.objects.create(emp_id=f"SLD-{last_number + 1}")
                else:
                    new_emp_id = EMP_ID.objects.create(emp_id="SLD-1")
                fm.emp_id=new_emp_id
                fm.save()
                messages.success(request, "User Added successfully.")
                return JsonResponse({'success': True, 'message': 'User created successfully!'})
            except ValidationError as e:
                # Handle explicit model-level validation errors
                return JsonResponse({'success': False, 'errors': {'non_field_errors': str(e)}}, status=400)
        else:
            # Handle form errors, including unique constraint violations
            errors = {
                field: [str(error) for error in error_list]
                for field, error_list in form.errors.items()
            }
            return JsonResponse({'success': False, 'errors': errors}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)

def update_user(request, id):
    user = get_object_or_404(CustomUser, id=id) 
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST, request.FILES, instance=user)  # Pre-populate the form with the user data
        if form.is_valid():
            fm=form.save(commit=False)  # Save the updated user data
            fm.save() 
            messages.success(request, "User details updated successfully.")
    else:
        form = UserRegistrationForm(instance=user)  # Pre-populate the form with existing data for GET request
    return render(request, 'admin_update_user.html', {'form': form, 'user': user})

from django.db.models import ProtectedError

def delete_user(request, id):
    try:
        user = get_object_or_404(CustomUser, id=id)

        # Prevent admin from deleting themselves
        if request.user == user:
            messages.error(request, "You cannot delete your own account.")
            return redirect('/admin/user-management/')  # Redirect to user list or another appropriate page

        if user:
            user=user.delete() 
            messages.success(request, "User deleted successfully.")
    
    except ProtectedError:
        messages.error(request, "Cannot delete user as they have related records.")
    
    except IntegrityError:
        messages.error(request, "Integrity constraint error while deleting user.")
    
    except DatabaseError as db_err:
        messages.error(request, f"Database error: {db_err}")
    
    except Exception as e:
        messages.error(request, f"Error: {e}")

    return redirect('/admin/user-management/')  # Redirect to the user list
 
@admin_required
def drivers_list(request):
    drivers = Driver.objects.select_related('user').all()  
    # Driver.objects.all().delete()
    form = DriverRegistrationForm()
    return render(request, "admin_drivers_list.html", {'form': form, 'drivers': drivers})


def create_driver(request):
    if request.method == 'POST':
        form = DriverRegistrationForm(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                with transaction.atomic():

                    form.save()
                    return JsonResponse({'success': True, 'message': 'Driver created successfully!'})
            except ValidationError as e:
                # Handle explicit model-level validation errors
                return JsonResponse({'success': False, 'errors': {'non_field_errors': str(e)}}, status=400)
            except Exception as e:
                # Catch any unexpected errors and return a 500 response
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
        else:
            # Handle form errors
            errors = {
                field: [str(error) for error in error_list]
                for field, error_list in form.errors.items()
            }
            return JsonResponse({'success': False, 'errors': errors}, status=400)
    
    # If the request method is not POST, return a method not allowed response
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)



def update_driver(request, id):
    driver = get_object_or_404(Driver, id=id)  # Safely retrieve the Driver instance or return a 404 error
    try:
        if request.method == 'POST':
            form = DriverUpdateForm(request.POST, request.FILES, instance=driver)  # Populate the form with the instance data
            if form.is_valid():
                is_active = form.cleaned_data['is_active']
                form.save()  # Save the updated driver instance
                
                # Update the related user object
                driver.user.is_active = is_active
                driver.user.save()

                messages.success(request, 'Driver Updated Successfully.')
        else:
            form = DriverUpdateForm(instance=driver)  # Populate the form with the existing driver data on GET request
    except IntegrityError:
        messages.error(request, "Integrity error while updating the driver.")
    except DatabaseError:
        messages.error(request, "Database error occurred. Please try again.")
    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")

    return render(request, 'admin_update_driver.html', {'form': form, 'driver': driver})


def delete_driver(request, id):
    driver = get_object_or_404(Driver, id=id)
    if driver:
        driver.delete()
        messages.success(request, 'Driver deleted successfully.')
    return redirect('/admin/drivers-list')


from django.db import IntegrityError

def import_drivers(request):
    
    if request.method == "POST":
        excel_file = request.FILES.get('Driver_file')
        password = make_password('Pass@123')
        if excel_file:
            try:
                # Load the Excel file
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active
                
                # List to hold successfully created Driver instances
                data_to_insert = []
                skipped_rows = []  # To track skipped rows with reasons

                # Iterate over each row in the Excel sheet (skipping header row)
                for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                    driver_name = row[0]  # Driver Name (index 0)
                    adhaar_number = row[1] if row[1] else None  # Aadhaar Number (index 1)
                    license_number = row[2] if row[2] else None  # License Number (index 2)
                    date_of_birth = row[3] if row[3] else None  # Date of Birth (index 3)
                    mobile_number = row[4] if row[4] else None  # Mobile Number (index 4)
                    date_joined = row[5] if row[5] else None  # Date Joined (index 5)

                    # Validate essential fields
                    if not driver_name or not (adhaar_number or mobile_number):
                        skipped_rows.append(
                            f"Row {row_index}: Missing essential fields (Driver Name: {driver_name}, Aadhaar: {adhaar_number}, Mobile: {mobile_number})."
                        )
                        continue

                    # Check if Aadhaar or Mobile number already exists
                    if adhaar_number and CustomUser.objects.filter(username=adhaar_number).exists():
                        skipped_rows.append(f"Row {row_index}: Aadhaar number {adhaar_number} is already registered.")
                        continue

                    if not adhaar_number and CustomUser.objects.filter(username=mobile_number).exists():
                        skipped_rows.append(f"Row {row_index}: Mobile number {mobile_number} is already registered.")
                        continue

                    # Create the EMP_ID instance if it doesn't exist
                    last_emp_id = EMP_ID.objects.order_by('-id').first()
                    if last_emp_id:
                        last_number = int(last_emp_id.emp_id.split('-')[1])
                        new_emp_id = EMP_ID.objects.create(emp_id=f"SLD-{last_number + 1}")
                    else:
                        new_emp_id = EMP_ID.objects.create(emp_id="SLD-1")

                    # Create or update the user (CustomUser)
                    username = mobile_number
                    user, created = CustomUser.objects.get_or_create(
                        username=username,
                        defaults={
                            'is_driver': True,
                            'password': password,
                            'emp_id': new_emp_id
                        }
                    )

                    if not created:
                        if user.is_driver:
                            skipped_rows.append(f"Row {row_index}: User {username} already exists as a driver.")
                            continue
                        user.is_driver = True
                        user.save()

                    # Create the Driver instance
                    driver = Driver(
                        user=user,
                        driver_name=driver_name,
                        license_number=license_number,
                        adhaar_number=adhaar_number,
                        mobile_number=mobile_number,
                        date_of_birth=date_of_birth,
                        date_joined=date_joined
                    )

                    try:
                        driver.save()  # Save the driver instance
                        data_to_insert.append(driver)  # Add to the list of successfully created records
                    except IntegrityError as e:
                        skipped_rows.append(
                            f"Row {row_index}: UNIQUE constraint failed (License Number: {license_number})."
                        )
                        continue  # Skip this row and continue to the next

                # Provide feedback to the user
                if data_to_insert:
                    messages.success(request, 'Drivers data imported and updated successfully.')
                else:
                    messages.error(request, 'No data imported. Check skipped rows for errors.')

                # Debugging: Log skipped rows
                if skipped_rows:
                    print("Skipped rows with reasons:")
                    for reason in skipped_rows:
                        print(reason)
                    messages.error(request, f"Skipped rows: {len(skipped_rows)}. Check logs for details.")

            except Exception as e:
                print(f"Error occurred during import: {str(e)}")
                messages.error(request, f'Error occurred during import: {str(e)}')
        else:
            messages.error(request, 'No file selected.')

        return redirect('/admin/drivers-list')
    return redirect('/admin/drivers-list')


# vehical data start
@admin_required
def vehicle_list(request):
    vehicle = Vehicle.objects.select_related('model_name')
    form = VehicleForm()  # Pass the form for vehicle creation
    return render(request, "admin_vehicle_list.html", {'vehicle': vehicle, 'form': form})

def create_vehicle(request):
    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                fm=form.save()
                messages.success(request,"Vehiccle created successfully!")
                return JsonResponse({'success': True, 'message': 'Vehiccle created successfully!'})
            except ValidationError as e:
                # Handle explicit model-level validation errors
                return JsonResponse({'success': False, 'errors': {'non_field_errors': str(e)}}, status=400)
        else:
            # Handle form errors, including unique constraint violations
            errors = {
                field: [str(error) for error in error_list]
                for field, error_list in form.errors.items()
            }
            return JsonResponse({'success': False, 'errors': errors}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
 

def update_vehicle(request, id):
    try:
        vehicle = get_object_or_404(Vehicle, id=id)  # Safely retrieve the Vehicle instance or return a 404 error

        if request.method == 'POST':
            form = VehicleForm(request.POST, instance=vehicle)  # Populate the form with the instance data
            if form.is_valid():
                form.save()  # Save the updated vehicle instance
                messages.success(request, "Vehicle updated successfully.")
        else:
            form = VehicleForm(instance=vehicle)  # Populate the form with existing vehicle data for GET request
    except IntegrityError:
        messages.error(request, "Integrity error while updating the vehicle.")
    except DatabaseError:
        messages.error(request, "Database error occurred. Please try again.")
    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")

    return render(request, 'admin_update_vehicle.html', {'form': form})

def import_vehicles(request):
    if request.method == "POST":
        excel_file = request.FILES.get('vehicle_file')
        if excel_file:
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active
                data_to_insert = []

                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    vehicle_number = row[0]
                    model_name = row[1]
                    owner_name = row[2] if len(row) > 2 else None  # Optional Owner Name

                    if not vehicle_number or not model_name:
                        continue  # Skip incomplete rows

                    # Get or Create the VehicleModel
                    vehicle_model, _ = VehicleModel.objects.get_or_create(model_name=model_name)

                    # Skip if Vehicle already exists
                    if Vehicle.objects.select_related('model_name').filter(vehicle_number=vehicle_number).exists():
                        messages.warning(request, f"Vehicle with number {vehicle_number} already exists. Skipping.")
                        continue

                    # Create Vehicle Object
                    new_vehicle = Vehicle(
                        vehicle_number=vehicle_number,
                        model_name=vehicle_model,
                        owner_name=owner_name
                    )
                    data_to_insert.append(new_vehicle)
                # Bulk Insert Vehicles
                if data_to_insert:
                    Vehicle.objects.bulk_create(data_to_insert)
                    messages.success(request, f'Data Imported Successfully. {len(data_to_insert)} vehicles added.')
                else:
                    messages.info(request, 'No new vehicles were added.')
            except Exception as e:
                messages.error(request, f'Error occurred during import: {str(e)}')
        else:
            messages.error(request, 'No file selected.')
        return redirect('/admin/vehicle-list')
    return redirect('/admin/vehicle-list')


def delete_vehicle(request, id):
    vehicle = get_object_or_404(Vehicle, id=id)
    if vehicle:
        vehicle.delete()
        messages.success(request, 'Vehicle deleted successfully.')
    return redirect('/admin/vehicle-list')



@admin_required
def technician_list(request):
    technicians=Technician.objects.select_related('emp_id')
    form = TechnicianRegistrationForm()
    return render(request, "admin_technician_list.html",{'form':form,'technicians':technicians})

def create_technician(request):
    if request.method == 'POST':
        form = TechnicianRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                fm=form.save(commit=False)
                last_emp_id = EMP_ID.objects.order_by('-id').first()
                if last_emp_id:
                    last_number = int(last_emp_id.emp_id.split('-')[1])
                    new_emp_id = EMP_ID.objects.create(emp_id=f"SLD-{last_number + 1}")
                else:
                    new_emp_id = EMP_ID.objects.create(emp_id="SLD-1")
                fm.emp_id=new_emp_id
                fm.save()
                
                messages.success(request,"Technician created successfully!")
                return JsonResponse({'success': True, 'message': 'Technician created successfully!'})
            except ValidationError as e:
                # Handle explicit model-level validation errors
                return JsonResponse({'success': False, 'errors': {'non_field_errors': str(e)}}, status=400)
        else:
            # Handle form errors, including unique constraint violations
            errors = {
                field: [str(error) for error in error_list]
                for field, error_list in form.errors.items()
            }
            return JsonResponse({'success': False, 'errors': errors}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
 

 
 
def update_technician(request, id):
    data = get_object_or_404(Technician, id=id)
    if request.method == 'POST':
        form = TechnicianUpdateForm(request.POST,request.FILES, instance=data)  # Pre-populate the form with the user data
        if form.is_valid():
            form.save()  # Save the updated user data
            messages.success(request, "Technician updated successfully.")
    else:
        form = TechnicianUpdateForm(instance=data)  # Pre-populate the form with existing data for GET request
    return render(request, 'admin_update_technician.html', {'form': form})


def delete_technician(request, id):
    technician = get_object_or_404(Technician, id=id)
    if technician:
        technician.delete()
        messages.success(request, 'Technician deleted successfully.')
    return redirect('/admin/technician-list')






@admin_required
def party_list(request):
    partys=Party.objects.all()
    form = PartyForm()
    return render(request, "admin_party_list.html",{'form':form,'partys':partys})

def create_party(request):
    if request.method == 'POST':
        form = PartyForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                fm=form.save()
                messages.success(request,"Party created successfully!")
                return JsonResponse({'success': True, 'message': 'Party created successfully!'})
            except ValidationError as e:
                # Handle explicit model-level validation errors
                return JsonResponse({'success': False, 'errors': {'non_field_errors': str(e)}}, status=400)
        else:
            # Handle form errors, including unique constraint violations
            errors = {
                field: [str(error) for error in error_list]
                for field, error_list in form.errors.items()
            }
            return JsonResponse({'success': False, 'errors': errors}, status=400)
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
 



def update_party(request, id):
    data = get_object_or_404(Party, id=id)
    if request.method == 'POST':
        form = PartyForm(request.POST, instance=data)  # Pre-populate the form with the user data
        if form.is_valid():
            form.save()  # Save the updated user data
            messages.success(request, "Party updated successfully.")
    else:
        form = PartyForm(instance=data)  # Pre-populate the form with existing data for GET request
    return render(request, 'admin_update_party.html', {'form': form})



def delete_party(request, id):
    party = get_object_or_404(Party, id=id)
    if party:
        party.delete()
        messages.success(request, 'Party deleted successfully.')
    return redirect('/admin/party-list')


from django.db.models import Sum
from datetime import datetime, timedelta

@admin_required
def vehicle_dashboard(request):
    vehicle_data = Vehicle.objects.select_related('model_name').filter(status='active')
    vehicle_number = request.GET.get('vehicle_number', '').strip()
    vehicle = None
    job_cards = []
    total_labour_cost = 0
    total_item_cost = 0
    grand_total_cost = 0
    
    emi_details=None 
    one_emi_amount=0
    remaining_emi_amount=0

    vehicle = Vehicle.objects.select_related('model_name').filter(vehicle_number__iexact=vehicle_number).first()
    try:
        if vehicle_number:
            # Get the vehicle by its vehicle_number
            
            if vehicle:
                # Query job cards related to this vehicle
                job_cards = JobCard.objects.filter(vehicle=vehicle)
                
                for job_card in job_cards:
                    # Aggregate total item cost for the current job card
                    items = JobCardItem.objects.filter(job_card=job_card)
                    item_cost = items.aggregate(Sum('total_cost'))['total_cost__sum'] or 0
                    
                    # Labour cost for the current job card
                    labour_cost = job_card.labour_cost or 0
                    
                    # Update totals
                    total_labour_cost += labour_cost
                    total_item_cost += item_cost

                # Grand total cost (labour + items)
                grand_total_cost = total_labour_cost + total_item_cost

                # Fetch EMI related to the vehicle
                emi_details = EMI.objects.filter(vehicle=vehicle).first() 

                if emi_details:
                    one_emi_amount = emi_details.loan_amount / emi_details.total_installments
                    remaining_emi_amount=int(one_emi_amount)*int(emi_details.remaining_installments)

    except Exception as e:
        # Log the exception if needed (e.g., with logging)
        return render(request, '404.html', {})

    context = {
        'vehicle': vehicle,
        'job_card_count': len(job_cards),
        'total_labour_cost': int(total_labour_cost),
        'total_item_cost': int(total_item_cost),
        'grand_total_cost': int(grand_total_cost),
        'job_cards': job_cards,
        'vehicle_data': vehicle_data,
        'is_vehicle_data': vehicle_number,
        
        # EMI details to be displayed
        'emi_details': emi_details, 
        'one_emi_amount':one_emi_amount,
        'remaining_emi_amount':remaining_emi_amount
    }
    return render(request, 'admin_vehicle_dashboard.html', context)


import qrcode
from io import BytesIO
from django.http import HttpResponse

# def generate_and_download_qr(request, vehicle_number):
#     """
#     Generate a QR code for the given vehicle name and download it directly.
#     """
#     # Create a QR code object
#     qr = qrcode.QRCode(
#         version=1,  # Controls the size of the QR Code
#         error_correction=qrcode.constants.ERROR_CORRECT_L,
#         box_size=10,
#         border=4,
#     )
    
#     # Add data to the QR code
#     qr.add_data(vehicle_number)
#     qr.make(fit=True)

#     # Create an image of the QR code
#     img = qr.make_image(fill_color="black", back_color="white")

#     # Save the QR code image to a buffer
#     buffer = BytesIO()
#     img.save(buffer, format="PNG")
#     buffer.seek(0)

#     # Create the HTTP response for direct download
#     response = HttpResponse(buffer, content_type="image/png")
#     response['Content-Disposition'] = f'attachment; filename="{vehicle_number}_QR.png"' 
#     return response

from django.shortcuts import render
from django.http import HttpResponse
import qrcode
from io import BytesIO
import base64

def generate_qr(request, vehicle_number):
    """
    Generate a QR code for the given vehicle number and display it on an HTML page.
    """
    vehicle=Vehicle.objects.get(vehicle_number=vehicle_number)
    # Create a QR code object
    qr = qrcode.QRCode(
        version=1,  # Controls the size of the QR Code
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    
    # Add data to the QR code
    qr.add_data(vehicle_number)
    qr.make(fit=True)

    # Create an image of the QR code
    img = qr.make_image(fill_color="black", back_color="white")

    # Save the QR code image to a buffer
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)

    # Convert the image to base64 string
    qr_image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

    # Render the HTML template and pass the base64 image string
    return render(request, 'admin_print_qr.html', {'qr_image_base64': qr_image_base64, 'vehicle': vehicle})







from collections import Counter
from datetime import date

def finance_dashboard(request):
    today = date.today()
    thirty_days_date = today + timedelta(days=30)
    active_vehicles_filter = Q(vehicle__status='active')
    
    thirty_days_record = {
        'policy_dues': Policy.objects.filter(active_vehicles_filter, due_date__range=[today, thirty_days_date]),
        'emi_dues': EMI.objects.filter(active_vehicles_filter, status='pending', next_due_date__range=[today, thirty_days_date]),
        'tax_dues': OtherDues.objects.filter(active_vehicles_filter, tax_due_date__range=[today, thirty_days_date]),
        'fitness_dues': OtherDues.objects.filter(active_vehicles_filter, fitness_due_date__range=[today, thirty_days_date]),
        'permit_dues': OtherDues.objects.filter(active_vehicles_filter, permit_due_date__range=[today, thirty_days_date]),
        'puc_dues': OtherDues.objects.filter(active_vehicles_filter, puc_due_date__range=[today, thirty_days_date]),
    }

    # Count the records efficiently using annotations
    thirty_days_counts = {
        key: value.aggregate(count=Count('id'))['count'] for key, value in thirty_days_record.items()
    }

    # Queries for past due dates using Q objects
    expire_dues_counts = Counter({
        "policy_dues": Policy.objects.filter(active_vehicles_filter, due_date__lt=today).exclude(due_date__isnull=True).count(),
        "emi_dues": EMI.objects.filter(active_vehicles_filter, status='pending', next_due_date__lt=today).exclude(next_due_date__isnull=True).count(),
        "tax_dues": OtherDues.objects.filter(active_vehicles_filter, tax_due_date__lt=today).exclude(tax_due_date__isnull=True).count(),
        "fitness_dues": OtherDues.objects.filter(active_vehicles_filter, fitness_due_date__lt=today).exclude(fitness_due_date__isnull=True).count(),
        "permit_dues": OtherDues.objects.filter(active_vehicles_filter, permit_due_date__lt=today).exclude(permit_due_date__isnull=True).count(),
        "puc_dues": OtherDues.objects.filter(active_vehicles_filter, puc_due_date__lt=today).exclude(puc_due_date__isnull=True).count(),
    })

    three_days_later = today + timedelta(days=3)
    # Query the different dues in the next 2 days
    emi_dues = EMI.objects.filter(next_due_date__range=[today, three_days_later],vehicle__status='active', status='pending').values('vehicle__id', 'vehicle__vehicle_number', 'next_due_date')
    policy_dues = Policy.objects.filter(due_date__range=[today, three_days_later],vehicle__status='active').values('vehicle__id', 'vehicle__vehicle_number', 'due_date')
    tax_dues = OtherDues.objects.filter(tax_due_date__range=[today, three_days_later],vehicle__status='active').values('vehicle__id', 'vehicle__vehicle_number', 'tax_due_date')
    fitness_dues = OtherDues.objects.filter(fitness_due_date__range=[today, three_days_later],vehicle__status='active').values('vehicle__id', 'vehicle__vehicle_number', 'fitness_due_date')
    permit_dues = OtherDues.objects.filter(permit_due_date__range=[today, three_days_later],vehicle__status='active').values('vehicle__id', 'vehicle__vehicle_number', 'permit_due_date')
    puc_dues = OtherDues.objects.filter(puc_due_date__range=[today, three_days_later],vehicle__status='active').values('vehicle__id', 'vehicle__vehicle_number', 'puc_due_date')

    # Add the due type as a custom field for each query result
    emi_dues_2 = [{'vehicle_id': due['vehicle__id'], 'vehicle_number': due['vehicle__vehicle_number'], 'due_date': due['next_due_date'], 'due_name': 'EMI'} for due in emi_dues]
    policy_dues_2 = [{'vehicle_id': due['vehicle__id'], 'vehicle_number': due['vehicle__vehicle_number'], 'due_date': due['due_date'], 'due_name': 'Policy'} for due in policy_dues]
    tax_dues_2 = [{'vehicle_id': due['vehicle__id'], 'vehicle_number': due['vehicle__vehicle_number'], 'due_date': due['tax_due_date'], 'due_name': 'Tax'} for due in tax_dues]
    fitness_dues_2 = [{'vehicle_id': due['vehicle__id'], 'vehicle_number': due['vehicle__vehicle_number'], 'due_date': due['fitness_due_date'], 'due_name': 'Fitness'} for due in fitness_dues]
    permit_dues_2 = [{'vehicle_id': due['vehicle__id'], 'vehicle_number': due['vehicle__vehicle_number'], 'due_date': due['permit_due_date'], 'due_name': 'Permit'} for due in permit_dues]
    puc_dues_2 = [{'vehicle_id': due['vehicle__id'], 'vehicle_number': due['vehicle__vehicle_number'], 'due_date': due['puc_due_date'], 'due_name': 'PUC'} for due in puc_dues]

    # Combine all dues data
    three_days_later_dues = emi_dues_2 + policy_dues_2 + tax_dues_2 + fitness_dues_2 + permit_dues_2 + puc_dues_2
 
    fifteen_days_ago = today - timedelta(days=15)
    one_days_ago = today - timedelta(days=1)

    emi_dues = EMI.objects.filter(next_due_date__range=[fifteen_days_ago, one_days_ago],vehicle__status='active', status='pending').values('vehicle__id', 'vehicle__vehicle_number', 'next_due_date')
    policy_dues = Policy.objects.filter(due_date__range=[fifteen_days_ago, one_days_ago],vehicle__status='active').values('vehicle__id', 'vehicle__vehicle_number', 'due_date')
    tax_dues = OtherDues.objects.filter(tax_due_date__range=[fifteen_days_ago, one_days_ago],vehicle__status='active').values('vehicle__id', 'vehicle__vehicle_number', 'tax_due_date')
    fitness_dues = OtherDues.objects.filter(fitness_due_date__range=[fifteen_days_ago, one_days_ago],vehicle__status='active').values('vehicle__id', 'vehicle__vehicle_number', 'fitness_due_date')
    permit_dues = OtherDues.objects.filter(permit_due_date__range=[fifteen_days_ago, one_days_ago],vehicle__status='active').values('vehicle__id', 'vehicle__vehicle_number', 'permit_due_date')
    puc_dues = OtherDues.objects.filter(puc_due_date__range=[fifteen_days_ago, one_days_ago],vehicle__status='active').values('vehicle__id', 'vehicle__vehicle_number', 'puc_due_date')


    emi_dues = [{'vehicle_id': due['vehicle__id'], 'vehicle_number': due['vehicle__vehicle_number'], 'due_date': due['next_due_date'], 'due_name': 'EMI'} for due in emi_dues]
    policy_dues = [{'vehicle_id': due['vehicle__id'], 'vehicle_number': due['vehicle__vehicle_number'], 'due_date': due['due_date'], 'due_name': 'Policy'} for due in policy_dues]
    tax_dues = [{'vehicle_id': due['vehicle__id'], 'vehicle_number': due['vehicle__vehicle_number'], 'due_date': due['tax_due_date'], 'due_name': 'Tax'} for due in tax_dues]
    fitness_dues = [{'vehicle_id': due['vehicle__id'], 'vehicle_number': due['vehicle__vehicle_number'], 'due_date': due['fitness_due_date'], 'due_name': 'Fitness'} for due in fitness_dues]
    permit_dues = [{'vehicle_id': due['vehicle__id'], 'vehicle_number': due['vehicle__vehicle_number'], 'due_date': due['permit_due_date'], 'due_name': 'Permit'} for due in permit_dues]
    puc_dues = [{'vehicle_id': due['vehicle__id'], 'vehicle_number': due['vehicle__vehicle_number'], 'due_date': due['puc_due_date'], 'due_name': 'PUC'} for due in puc_dues]

    # Combine all dues data
    fifteen_days_ago_expire_dues = emi_dues + policy_dues + tax_dues + fitness_dues + permit_dues + puc_dues
  
    # Pass data to template
    return render(request, 'admin_finance_dashboard.html', { 
        'thirty_days_counts':thirty_days_counts,
        'thirty_days_records':thirty_days_record,
        'expire_dues_counts':expire_dues_counts,
        'three_days_later_dues':three_days_later_dues,
        'fifteen_days_ago_expire_dues':fifteen_days_ago_expire_dues

    })
 
 

def finance_vehicle_list(request):

    # Instantiate the filter form
    filter_form = VehicleFilterForFinance(request.GET or None)
    queryset = Vehicle.objects.select_related('model_name').filter(status='active')

    # Initialize filters
    start_date = filter_form.cleaned_data.get('start_date') if filter_form.is_valid() else None
    end_date = filter_form.cleaned_data.get('end_date') if filter_form.is_valid() else None

    # Prepare the context
    context = []
    for vehicle in queryset:
        # Get related data for each vehicle
        policies = vehicle.policies.select_related()
        emis = vehicle.emis.select_related()
        other_dues = vehicle.otherdues_set.first()   

        # Extract relevant due dates
        policy_due_date = policies[0].due_date if policies else None
        emi_due_date = emis[0].next_due_date if emis else None
        tax_due_date = other_dues.tax_due_date if other_dues else None
        fitness_due_date = other_dues.fitness_due_date if other_dues else None
        permit_due_date = other_dues.permit_due_date if other_dues else None
        puc_due_date = other_dues.puc_due_date if other_dues else None

        # Apply date filtering
        if start_date and not end_date:
            # Only dates greater than or equal to start_date
            if not (
                (policy_due_date and policy_due_date >= start_date) or
                (emi_due_date and emi_due_date >= start_date) or 
                (tax_due_date and tax_due_date >= start_date) or
                (fitness_due_date and fitness_due_date >= start_date) or
                (permit_due_date and permit_due_date >= start_date) or
                (puc_due_date and puc_due_date >= start_date)
            ):
                continue  # Skip vehicles that don't match the filter
        elif end_date and not start_date:
            # Only dates less than or equal to end_date
            if not (
                (policy_due_date and policy_due_date <= end_date) or
                (emi_due_date and emi_due_date <= end_date) or 
                (tax_due_date and tax_due_date <= end_date) or
                (fitness_due_date and fitness_due_date <= end_date) or
                (permit_due_date and permit_due_date <= end_date) or
                (puc_due_date and puc_due_date <= end_date)
            ):
                continue  # Skip vehicles that don't match the filter
        elif start_date and end_date:
            # Dates within the range of start_date and end_date
            if not (
                (policy_due_date and start_date <= policy_due_date <= end_date) or
                (emi_due_date and start_date <= emi_due_date <= end_date) or 
                (tax_due_date and start_date <= tax_due_date <= end_date) or
                (fitness_due_date and start_date <= fitness_due_date <= end_date) or
                (permit_due_date and start_date <= permit_due_date <= end_date) or
                (puc_due_date and start_date <= puc_due_date <= end_date)
            ):
                continue  # Skip vehicles that don't match the filter

        # Add vehicle info to the context
        context.append({
            'id': vehicle.id,
            'vehicle_number': vehicle.vehicle_number,
            'policy_due_date': policy_due_date,
            'emi_due_date': emi_due_date, 
            'tax_due_date': tax_due_date,
            'fitness_due_date': fitness_due_date,
            'permit_due_date': permit_due_date,
            'puc_due_date': puc_due_date,
        })

    # Render the template with the filter form and context
    return render(request, 'admin_finance_vehicle_list.html', {
        'filter_form': filter_form,
        'context': context
    })



def finance_vehicle_dashboard(request,id): 

    vehicle = Vehicle.objects.select_related('model_name') or None
    vehicle = Vehicle.objects.get(id=id)
    other_dues_data=OtherDues.objects.filter(vehicle=vehicle) 

    other_dues=OtherDues.objects.select_related('vehicle ').filter(vehicle=vehicle).first() 
    other_dues_id=0
    if other_dues:
        other_dues_id=other_dues.id 

    is_data_exist=other_dues_data.exists()

    emi=EMI.objects.filter(vehicle=vehicle).first() 
    policy=Policy.objects.filter(vehicle=vehicle).first()


    context={
        'vehicle_id':id,
        'other_dues_id':other_dues_id,
        'other_dues':other_dues,
        'vehicle':vehicle, 
        'is_data_exist':is_data_exist,
        'emi':emi,
        'policy':policy, 
    }
    return render(request, 'admin_finance_vehicle_dashboard.html',context)



